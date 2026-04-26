"""
main.py — Tata Motors Financial Dashboard Pipeline
Author: Bonthu Raj Dhanush
LinkedIn: linkedin.com/in/raj-dhanush-6a1010322

Run this file to execute the full pipeline:
1. Load CSV data into SQLite database
2. Run all SQL analytical queries
3. Generate all charts
4. Build Excel dashboard
"""

import os
import sys

print("=" * 60)
print("TATA MOTORS FINANCIAL DASHBOARD")
print("Author: Bonthu Raj Dhanush")
print("=" * 60)

# Step 1 — Database
print("\n[1/3] Setting up database and running SQL queries...")
from database import create_database, run_all_queries
df = create_database()
results = run_all_queries()

print("\n📊 Key Query Results:")
print("\n-- Revenue CAGR --")
print(results["cagr"].to_string(index=False))

print("\n-- PAT Turnaround --")
print(results["milestones"].to_string(index=False))

# Step 2 — Charts
print("\n[2/3] Generating charts...")
from visualise import generate_all_charts
generate_all_charts()

# Step 3 — Excel Dashboard
print("\n[3/3] Building Excel dashboard...")
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Dashboard"

DARK_BLUE = "1F3864"
MID_BLUE  = "2E75B6"
WHITE     = "FFFFFF"
GREEN_BG  = "E2EFDA"
GREEN_FG  = "375623"
RED_BG    = "FCE4D6"
RED_FG    = "9C0006"
GREY_BG   = "F2F2F2"

def border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def hdr(ws, r, c, text, bg=DARK_BLUE, fg=WHITE, bold=True, size=10):
    cell = ws.cell(row=r, column=c, value=text)
    cell.font = Font(name="Arial", bold=bold, color=fg, size=size)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border()
    return cell

def val(ws, r, c, value, fmt="#,##0.00", bold=False, bg=WHITE, fg="000000"):
    cell = ws.cell(row=r, column=c, value=value)
    cell.number_format = fmt
    cell.font = Font(name="Arial", bold=bold, color=fg, size=10)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.border = border()
    return cell

# Title
ws.merge_cells("A1:L1")
t = ws["A1"]
t.value = "TATA MOTORS LTD — Financial Performance Dashboard (FY2016–FY2025)"
t.font = Font(name="Arial", bold=True, size=14, color=WHITE)
t.fill = PatternFill("solid", fgColor=DARK_BLUE)
t.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 32

ws.merge_cells("A2:L2")
s = ws["A2"]
s.value = "Author: Bonthu Raj Dhanush  |  Data Source: Screener.in — Tata Motors Annual Reports  |  Tools: Python · SQL · Excel"
s.font = Font(name="Arial", italic=True, size=9, color="595959")
s.fill = PatternFill("solid", fgColor=GREY_BG)
s.alignment = Alignment(horizontal="center", vertical="center")

# KPI Cards
kpis = [
    ("Revenue FY2025", "₹4,86,483 Cr", "+10.7% YoY", GREEN_BG, GREEN_FG),
    ("EBITDA Margin", "9.35%", "vs 0.51% in FY2022", GREEN_BG, GREEN_FG),
    ("PAT FY2025", "₹27,391 Cr", "Turnaround from -₹34,153 Cr", GREEN_BG, GREEN_FG),
    ("EPS FY2025", "₹71.64", "vs -₹99.45 in FY2022", GREEN_BG, GREEN_FG),
    ("Revenue CAGR", "~6.1%", "FY2016 to FY2025", MID_BLUE, WHITE),
]

ws.merge_cells("A4:L4")
kh = ws["A4"]
kh.value = "KEY PERFORMANCE INDICATORS"
kh.font = Font(name="Arial", bold=True, size=11, color=WHITE)
kh.fill = PatternFill("solid", fgColor=MID_BLUE)
kh.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[4].height = 22

for ki, (name, kval, ksub, kbg, kfg) in enumerate(kpis):
    col = 1 + ki * 2
    for r in [5, 6, 7]:
        ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col+1)
        ws.row_dimensions[r].height = 20

    ws.cell(5, col, name).font = Font(name="Arial", bold=True, size=9, color="595959")
    ws.cell(5, col).fill = PatternFill("solid", fgColor=kbg)
    ws.cell(5, col).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(5, col).border = border()

    ws.cell(6, col, kval).font = Font(name="Arial", bold=True, size=12, color=kfg)
    ws.cell(6, col).fill = PatternFill("solid", fgColor=kbg)
    ws.cell(6, col).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(6, col).border = border()

    ws.cell(7, col, ksub).font = Font(name="Arial", size=8, color=kfg)
    ws.cell(7, col).fill = PatternFill("solid", fgColor=kbg)
    ws.cell(7, col).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(7, col).border = border()

# Main Data Table
ws.merge_cells("A9:L9")
th = ws["A9"]
th.value = "PROFIT & LOSS STATEMENT — 10 YEAR TREND (₹ CRORES)"
th.font = Font(name="Arial", bold=True, size=11, color=WHITE)
th.fill = PatternFill("solid", fgColor=MID_BLUE)
th.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[9].height = 22

cols_h = ["Year", "Revenue", "Expenses", "Op. Profit", "Other Inc.",
          "Depreciation", "Interest", "PBT", "Tax", "PAT", "EPS", "EBITDA %"]
for ci, h in enumerate(cols_h):
    hdr(ws, 10, ci+1, h)
ws.row_dimensions[10].height = 22

for ri, row in df.iterrows():
    r = 11 + ri
    is_profit = row["pat"] >= 0
    bg = WHITE if ri % 2 == 0 else GREY_BG

    ws.cell(r, 1, row["year"]).font = Font(name="Arial", bold=True, size=10)
    ws.cell(r, 1).fill = PatternFill("solid", fgColor=bg)
    ws.cell(r, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(r, 1).border = border()

    for ci, col in enumerate(["revenue", "expenses", "operating_profit",
                               "other_income", "depreciation", "interest",
                               "pbt", "tax", "pat"]):
        v = row[col]
        cell_bg = bg
        cell_fg = "000000"
        if col == "pat":
            cell_bg = GREEN_BG if is_profit else RED_BG
            cell_fg = GREEN_FG if is_profit else RED_FG
        val(ws, r, ci+2, v, "#,##0.00", col in ["pat", "operating_profit"],
            cell_bg, cell_fg)

    val(ws, r, 11, row["eps"], "#,##0.00", False, bg,
        GREEN_FG if is_profit else RED_FG)

    opm_cell = ws.cell(r, 12, row["opm_pct"] / 100)
    opm_cell.number_format = "0.00%"
    opm_cell.font = Font(name="Arial", size=10, bold=True,
                         color=GREEN_FG if row["opm_pct"] >= 0 else RED_FG)
    opm_cell.fill = PatternFill("solid", fgColor=GREEN_BG if row["opm_pct"] >= 0 else RED_BG)
    opm_cell.alignment = Alignment(horizontal="right", vertical="center")
    opm_cell.border = border()
    ws.row_dimensions[r].height = 18

# Column widths
widths = [8, 13, 13, 12, 10, 12, 10, 12, 10, 12, 10, 10]
for i, w in enumerate(widths):
    ws.column_dimensions[get_column_letter(i+1)].width = w

# Insights sheet
ws2 = wb.create_sheet("SQL Query Results")
ws2.merge_cells("A1:F1")
t2 = ws2["A1"]
t2.value = "SQL Analysis Results — Tata Motors Financial Dashboard"
t2.font = Font(name="Arial", bold=True, size=13, color=WHITE)
t2.fill = PatternFill("solid", fgColor=DARK_BLUE)
t2.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 28

row_offset = 3
for query_name, result_df in results.items():
    ws2.merge_cells(start_row=row_offset, start_column=1,
                    end_row=row_offset, end_column=len(result_df.columns))
    hc = ws2.cell(row_offset, 1, f"Query: {query_name.upper()}")
    hc.font = Font(name="Arial", bold=True, size=10, color=WHITE)
    hc.fill = PatternFill("solid", fgColor=MID_BLUE)
    hc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    hc.border = border()
    ws2.row_dimensions[row_offset].height = 20
    row_offset += 1

    for ci, col in enumerate(result_df.columns):
        hdr(ws2, row_offset, ci+1, col, bg=DARK_BLUE)
        ws2.column_dimensions[get_column_letter(ci+1)].width = 18
    ws2.row_dimensions[row_offset].height = 20
    row_offset += 1

    for _, drow in result_df.iterrows():
        for ci, v in enumerate(drow):
            c = ws2.cell(row_offset, ci+1, v)
            c.font = Font(name="Arial", size=9)
            c.border = border()
            c.alignment = Alignment(horizontal="right" if isinstance(v, (int, float)) else "left",
                                    vertical="center")
        ws2.row_dimensions[row_offset].height = 16
        row_offset += 1
    row_offset += 2

output_path = os.path.join(os.path.dirname(__file__), "outputs", "tata_motors_dashboard.xlsx")
wb.save(output_path)
print(f"Excel dashboard saved: {output_path}")

print("\n" + "=" * 60)
print("PIPELINE COMPLETE")
print(f"Charts: /charts/ ({5} files)")
print(f"Excel:  /outputs/tata_motors_dashboard.xlsx")
print(f"DB:     /data/tata_motors.db")
print("=" * 60)
